from __future__ import annotations

import csv
import io
import math
from collections import defaultdict
from typing import Any

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import HTMLResponse, Response
from openpyxl import Workbook, load_workbook

router = APIRouter()

REQUIRED_COLUMNS = [
    "employee_id",
    "snapshot_date",
    "department",
    "role_level",
    "tenure_months",
    "monthly_salary",
    "overtime_hours_30d",
    "avg_weekly_hours",
    "absent_days_90d",
    "late_days_90d",
    "performance_score",
    "engagement_score",
    "manager_changes_12m",
    "months_since_promotion",
    "training_hours_90d",
    "remote_days_week",
    "commute_minutes",
    "pay_raise_months_ago",
]

OPTIONAL_COLUMNS = ["left_company"]

BASELINE_LOGIT = -3.2
HIGH_RISK_THRESHOLD = 0.25
MEDIUM_RISK_THRESHOLD = 0.12

TEMPLATE_ROWS = [
    ["EMP-0001","2026-08-01","Engineering","Mid",18,5200,8,41,1,1,4.1,78,0,14,18,3,22,10,0],
    ["EMP-0002","2026-08-01","Sales","Senior",29,7600,24,49,4,3,3.7,44,1,31,6,2,55,22,0],
    ["EMP-0003","2026-08-01","Customer Success","Mid",11,5000,17,46,5,4,3.9,39,2,28,4,4,48,19,0],
    ["EMP-0004","2026-08-01","Finance","Lead",46,9100,5,40,0,0,4.4,83,0,8,22,2,18,6,0],
    ["EMP-0005","2026-08-01","Operations","Junior",5,3400,21,51,6,5,3.2,42,1,18,2,0,61,17,0],
    ["EMP-0006","2026-08-01","Engineering","Senior",37,7900,7,42,1,0,4.5,88,0,10,25,5,12,8,0],
    ["EMP-0007","2026-08-01","Sales","Mid",7,4800,29,54,3,4,3.5,36,2,40,3,1,52,26,1],
    ["EMP-0008","2026-08-01","Customer Success","Lead",33,8600,10,43,2,1,4.0,72,0,19,14,3,27,12,0],
]

SIGNAL_GUIDANCE = {
    "Low engagement": {
        "benchmark": "Engagement below 60/100",
        "why": (
            "Lower engagement can be associated with weaker attachment to the role or team. "
            "The model only raises this signal below 60, and the effect grows gradually as the score falls."
        ),
        "caution": (
            "Engagement is a directional workforce signal, not proof that an employee plans to leave. "
            "Use it to prompt better listening and retention conversations."
        ),
    },
    "High overtime": {
        "benchmark": "More than 12 overtime hours in 30 days",
        "why": (
            "Sustained overtime can indicate workload pressure or capacity imbalance. "
            "The model increases risk only for overtime above the 12-hour reference point."
        ),
        "caution": (
            "Overtime can be seasonal, voluntary, or project-specific. Review the operating context before interpreting it."
        ),
    },
    "Long weekly hours": {
        "benchmark": "Average above 45 hours/week",
        "why": (
            "Consistently long work weeks can be a workload-strain signal. "
            "Risk begins to rise only above the 45-hour reference point."
        ),
        "caution": (
            "This is a workload pattern, not a judgment about performance or commitment."
        ),
    },
    "Absence frequency": {
        "benchmark": "Each absent day in the last 90 days contributes modestly",
        "why": (
            "Repeated absence can coincide with scheduling friction or reduced continuity, so the model treats it as a weak contextual signal."
        ),
        "caution": (
            "Absence has many legitimate causes. Never infer health, disability, intent, or misconduct from this signal."
        ),
    },
    "Late arrivals": {
        "benchmark": "Each late arrival in the last 90 days contributes slightly",
        "why": (
            "Repeated lateness can indicate scheduling or commute friction, so the model uses it as a low-weight supporting signal."
        ),
        "caution": (
            "Treat this as operational context only. Do not use it as a disciplinary recommendation."
        ),
    },
    "Manager changes": {
        "benchmark": "Manager changes during the last 12 months",
        "why": (
            "Repeated manager changes can reduce continuity, clarity, and relationship stability. "
            "The model therefore gives this signal a larger stepwise effect."
        ),
        "caution": (
            "A manager change can also be positive. Review whether the transition improved or disrupted the employee experience."
        ),
    },
    "Time since promotion": {
        "benchmark": "More than 24 months since promotion",
        "why": (
            "Long periods without visible progression can create career-growth pressure. "
            "The model starts increasing risk after 24 months."
        ),
        "caution": (
            "Promotion cadence differs by role and company. Interpret against the employee's career path and expectations."
        ),
    },
    "Time since pay raise": {
        "benchmark": "More than 18 months since pay raise",
        "why": (
            "A long gap in compensation progression can contribute to retention pressure. "
            "The model starts increasing risk after 18 months."
        ),
        "caution": (
            "This is not a compensation recommendation and should not be used to determine pay."
        ),
    },
    "Long commute": {
        "benchmark": "More than 40 minutes one way",
        "why": (
            "Long commutes can add recurring time friction. The model begins increasing the signal beyond 40 minutes."
        ),
        "caution": (
            "Remote-work patterns and personal preferences vary. Use this only as optional retention context."
        ),
    },
    "Very short tenure": {
        "benchmark": "Less than 6 months tenure",
        "why": (
            "Early-tenure employees often have less organizational attachment and are still validating role fit, "
            "so the model adds a one-time early-tenure risk factor."
        ),
        "caution": (
            "Short tenure alone should never trigger an employment action. It is most useful for onboarding and support planning."
        ),
    },
    "Recent training": {
        "benchmark": "Training hours during the last 90 days",
        "why": (
            "Recent development activity is treated as a protective signal because learning investment can increase role growth and connection."
        ),
        "caution": (
            "Training does not guarantee retention; it simply offsets some modeled risk when present."
        ),
    },
}


def _float(row: dict[str, Any], key: str, default: float = 0.0) -> float:
    value = row.get(key, default)
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        raise ValueError(f"{key} must be numeric")


def _int01(value: Any) -> int:
    if value in (None, ""):
        return 0
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "left", "exited"}:
        return 1
    if text in {"0", "false", "no", "n", "active", "stayed"}:
        return 0
    try:
        return 1 if float(text) >= 1 else 0
    except ValueError:
        raise ValueError("left_company must be 0/1, true/false, active/left")


def _sigmoid(value: float) -> float:
    return 1.0 / (1.0 + math.exp(-value))


def employee_attrition_score(
    row: dict[str, Any],
) -> tuple[float, list[dict[str, Any]], dict[str, Any]]:
    engagement = _float(row, "engagement_score")
    overtime = _float(row, "overtime_hours_30d")
    weekly = _float(row, "avg_weekly_hours")
    absent = _float(row, "absent_days_90d")
    late = _float(row, "late_days_90d")
    manager_changes = _float(row, "manager_changes_12m")
    months_promotion = _float(row, "months_since_promotion")
    training = _float(row, "training_hours_90d")
    commute = _float(row, "commute_minutes")
    raise_months = _float(row, "pay_raise_months_ago")
    tenure = _float(row, "tenure_months")

    contribution_rows: list[dict[str, Any]] = []
    z = BASELINE_LOGIT

    def add(label: str, amount: float, value: float, display_value: str):
        nonlocal z
        z += amount
        if abs(amount) >= 0.05:
            contribution_rows.append(
                {
                    "signal": label,
                    "direction": "raises risk" if amount > 0 else "reduces risk",
                    "impact_log_odds": amount,
                    "value": value,
                    "display_value": display_value,
                }
            )

    add(
        "Low engagement",
        max(0.0, 60.0 - engagement) * 0.035,
        engagement,
        f"{engagement:.0f}/100 engagement",
    )
    add(
        "High overtime",
        max(0.0, overtime - 12.0) * 0.045,
        overtime,
        f"{overtime:.1f} overtime hours / 30d",
    )
    add(
        "Long weekly hours",
        max(0.0, weekly - 45.0) * 0.06,
        weekly,
        f"{weekly:.1f} average weekly hours",
    )
    add(
        "Absence frequency",
        absent * 0.12,
        absent,
        f"{absent:.0f} absent days / 90d",
    )
    add(
        "Late arrivals",
        late * 0.06,
        late,
        f"{late:.0f} late arrivals / 90d",
    )
    add(
        "Manager changes",
        manager_changes * 0.35,
        manager_changes,
        f"{manager_changes:.0f} manager changes / 12m",
    )
    add(
        "Time since promotion",
        max(0.0, months_promotion - 24.0) * 0.025,
        months_promotion,
        f"{months_promotion:.0f} months since promotion",
    )
    add(
        "Time since pay raise",
        max(0.0, raise_months - 18.0) * 0.025,
        raise_months,
        f"{raise_months:.0f} months since pay raise",
    )
    add(
        "Long commute",
        max(0.0, commute - 40.0) * 0.012,
        commute,
        f"{commute:.0f} minute commute",
    )
    add(
        "Very short tenure",
        0.45 if tenure < 6 else 0.0,
        tenure,
        f"{tenure:.0f} months tenure",
    )
    add(
        "Recent training",
        -min(training, 30.0) * 0.012,
        training,
        f"{training:.1f} training hours / 90d",
    )

    probability = _sigmoid(z)
    baseline_probability = _sigmoid(BASELINE_LOGIT)

    drivers = []
    for item in contribution_rows:
        amount = item["impact_log_odds"]
        counterfactual_probability = _sigmoid(z - amount)
        impact_pp = (probability - counterfactual_probability) * 100.0
        guidance = SIGNAL_GUIDANCE[item["signal"]]
        drivers.append(
            {
                "signal": item["signal"],
                "direction": item["direction"],
                "impact": round(amount, 3),
                "model_effect_pp": round(impact_pp, 2),
                "value": item["value"],
                "display_value": item["display_value"],
                "benchmark": guidance["benchmark"],
                "why": guidance["why"],
                "caution": guidance["caution"],
            }
        )

    drivers.sort(key=lambda item: abs(item["model_effect_pp"]), reverse=True)
    drivers = drivers[:5]

    upward = [item for item in drivers if item["direction"] == "raises risk"]
    protective = [item for item in drivers if item["direction"] == "reduces risk"]

    if upward:
        strongest = ", ".join(item["signal"].lower() for item in upward[:3])
        summary = (
            f"The model estimates {probability * 100:.1f}% attrition risk, "
            f"{max(0.0, (probability - baseline_probability) * 100):.1f} percentage points "
            f"above its neutral baseline. The strongest modeled pressure comes from {strongest}."
        )
    else:
        summary = (
            f"The model estimates {probability * 100:.1f}% attrition risk. "
            "No major upward operating signal is currently dominating the score."
        )

    if protective:
        summary += (
            f" {protective[0]['signal']} is currently offsetting part of that modeled pressure."
        )

    context = {
        "baseline_probability": round(baseline_probability, 4),
        "risk_above_baseline_pp": round((probability - baseline_probability) * 100.0, 2),
        "upward_signal_count": sum(
            1 for item in contribution_rows if item["impact_log_odds"] > 0
        ),
        "protective_signal_count": sum(
            1 for item in contribution_rows if item["impact_log_odds"] < 0
        ),
        "summary": summary,
    }

    return probability, drivers, context


def _normalize_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        raise HTTPException(
            status_code=400, detail="The uploaded file contains no employee rows."
        )

    headers = {str(k).strip() for k in rows[0].keys() if k is not None}
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise HTTPException(
            status_code=400,
            detail={"message": "Missing required columns", "missing_columns": missing},
        )

    latest: dict[str, dict[str, Any]] = {}
    for index, raw in enumerate(rows, start=2):
        row = {str(k).strip(): value for k, value in raw.items() if k is not None}
        employee_id = str(row.get("employee_id", "")).strip()
        if not employee_id:
            raise HTTPException(
                status_code=400, detail=f"Row {index}: employee_id is required"
            )

        snapshot = str(row.get("snapshot_date", "")).strip()
        current = latest.get(employee_id)
        if current is None or snapshot >= str(current.get("snapshot_date", "")):
            latest[employee_id] = row

    return list(latest.values())


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = (
        wb["Employee_Data"]
        if "Employee_Data" in wb.sheetnames
        else wb[wb.sheetnames[0]]
    )
    iterator = ws.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return []

    headers = [str(value).strip() if value is not None else "" for value in header_row]
    rows = []
    for values in iterator:
        if not any(value not in (None, "") for value in values):
            continue
        rows.append(
            {
                headers[index]: values[index]
                for index in range(min(len(headers), len(values)))
                if headers[index]
            }
        )
    return rows


def analyze_employee_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    normalized = _normalize_rows(rows)
    scored = []
    exits = 0
    active = 0
    risk_sum = 0.0
    high_risk = 0
    engagement_values = []
    by_department: dict[str, list[float]] = defaultdict(list)

    has_exit_column = any("left_company" in row for row in normalized)

    for row in normalized:
        exited = _int01(row.get("left_company", 0)) if has_exit_column else 0
        exits += exited

        try:
            engagement_values.append(_float(row, "engagement_score"))
        except ValueError as exc:
            raise HTTPException(
                status_code=400, detail=f"{row.get('employee_id')}: {exc}"
            )

        if exited:
            continue

        active += 1
        try:
            risk, drivers, context = employee_attrition_score(row)
        except ValueError as exc:
            raise HTTPException(
                status_code=400, detail=f"{row.get('employee_id')}: {exc}"
            )

        risk_sum += risk
        if risk >= HIGH_RISK_THRESHOLD:
            band = "High"
            high_risk += 1
        elif risk >= MEDIUM_RISK_THRESHOLD:
            band = "Medium"
        else:
            band = "Low"

        department = str(row.get("department", "Unknown"))
        by_department[department].append(risk)

        scored.append(
            {
                "employee_id": str(row["employee_id"]),
                "department": department,
                "role_level": str(row.get("role_level", "")),
                "attrition_probability": round(risk, 4),
                "risk_band": band,
                "engagement_score": _float(row, "engagement_score"),
                "overtime_hours_30d": _float(row, "overtime_hours_30d"),
                "top_signals": drivers,
                "model_context": context,
            }
        )

    scored.sort(key=lambda item: item["attrition_probability"], reverse=True)
    total = len(normalized)

    department_summary = [
        {
            "department": department,
            "active_employees": len(values),
            "predicted_attrition_rate": round(sum(values) / len(values), 4),
        }
        for department, values in by_department.items()
    ]
    department_summary.sort(
        key=lambda item: item["predicted_attrition_rate"], reverse=True
    )

    baseline_probability = round(_sigmoid(BASELINE_LOGIT), 4)

    return {
        "kpis": {
            "unique_employees": total,
            "active_employees": active,
            "historical_exits": exits if has_exit_column else None,
            "observed_churn_rate": (
                round(exits / total, 4) if has_exit_column and total else None
            ),
            "predicted_attrition_rate": (
                round(risk_sum / active, 4) if active else 0.0
            ),
            "high_risk_active_employees": high_risk,
            "average_engagement_score": (
                round(sum(engagement_values) / len(engagement_values), 1)
                if engagement_values
                else None
            ),
        },
        "department_summary": department_summary,
        "employees": scored[:200],
        "methodology": {
            "model_type": "Transparent logistic attrition-risk model",
            "baseline_probability": baseline_probability,
            "medium_risk_threshold": MEDIUM_RISK_THRESHOLD,
            "high_risk_threshold": HIGH_RISK_THRESHOLD,
            "scope": (
                "Retention and workforce-planning signal review. "
                "The model combines operational indicators into a probability-like risk score."
            ),
            "interpretation": (
                "A raised signal means the uploaded value crossed a model reference point and "
                "increased the score. It does not prove why an employee might leave."
            ),
        },
        "model_note": (
            "This demo is designed for aggregate retention and workforce-planning analysis. "
            "It must not be used to make termination, hiring, compensation, promotion, discipline, "
            "or other employment decisions about an individual."
        ),
    }


def _csv_template() -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(REQUIRED_COLUMNS + OPTIONAL_COLUMNS)
    writer.writerows(TEMPLATE_ROWS)
    return output.getvalue().encode("utf-8")


def _xlsx_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee_Data"
    ws.append(REQUIRED_COLUMNS + OPTIONAL_COLUMNS)
    for row in TEMPLATE_ROWS:
        ws.append(row)
    ws.freeze_panes = "A2"

    widths = {
        "A": 14, "B": 14, "C": 20, "D": 12, "E": 14, "F": 15, "G": 19,
        "H": 17, "I": 18, "J": 16, "K": 18, "L": 18, "M": 20, "N": 22,
        "O": 20, "P": 18, "Q": 16, "R": 20, "S": 14,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    dd = wb.create_sheet("Data_Dictionary")
    dd.append(["Column", "Required", "Description"])
    descriptions = {
        "employee_id": "Unique employee identifier; avoid names/emails in demo uploads.",
        "snapshot_date": "Snapshot date in YYYY-MM-DD format.",
        "department": "Department or business function.",
        "role_level": "Role level such as Junior, Mid, Senior, Lead, Manager.",
        "tenure_months": "Months employed at snapshot date.",
        "monthly_salary": "Monthly base salary; not used by the demo risk score.",
        "overtime_hours_30d": "Overtime hours in the last 30 days.",
        "avg_weekly_hours": "Average weekly working hours.",
        "absent_days_90d": "Absent days in the last 90 days.",
        "late_days_90d": "Late arrivals in the last 90 days.",
        "performance_score": "1-5 performance score; not used by the demo risk score.",
        "engagement_score": "0-100 engagement score.",
        "manager_changes_12m": "Number of manager changes in the last 12 months.",
        "months_since_promotion": "Months since last promotion.",
        "training_hours_90d": "Training hours in the last 90 days.",
        "remote_days_week": "Average remote days per week; not used by the demo risk score.",
        "commute_minutes": "One-way commute minutes.",
        "pay_raise_months_ago": "Months since last pay raise.",
        "left_company": (
            "Optional historical outcome: 1 exited, 0 stayed. "
            "Used for observed churn only."
        ),
    }
    for column in REQUIRED_COLUMNS + OPTIONAL_COLUMNS:
        dd.append(
            [
                column,
                "Optional" if column in OPTIONAL_COLUMNS else "Yes",
                descriptions[column],
            ]
        )

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


@router.get("/employees", response_class=HTMLResponse)
def employee_page():
    return EMPLOYEE_PAGE


@router.get("/api/employee-attrition/template.csv")
def download_csv_template():
    return Response(
        content=_csv_template(),
        media_type="text/csv",
        headers={
            "Content-Disposition": 'attachment; filename="employee_attrition_template.csv"'
        },
    )


@router.get("/api/employee-attrition/template.xlsx")
def download_xlsx_template():
    return Response(
        content=_xlsx_template(),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": 'attachment; filename="employee_attrition_template.xlsx"'
        },
    )


@router.post("/api/employee-attrition/upload")
async def upload_employee_file(file: UploadFile = File(...)):
    filename = (file.filename or "").lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Upload a .csv or .xlsx file.")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Demo upload limit is 5 MB.")

    try:
        rows = (
            _parse_csv(content)
            if filename.endswith(".csv")
            else _parse_xlsx(content)
        )
    except Exception as exc:
        raise HTTPException(
            status_code=400, detail=f"Could not parse file: {exc}"
        ) from exc

    if len(rows) > 5000:
        raise HTTPException(
            status_code=413, detail="Demo upload limit is 5,000 rows."
        )

    return analyze_employee_rows(rows)


EMPLOYEE_PAGE = """<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Employee Attrition Analyzer — Technyder</title>
<style>
*{box-sizing:border-box}
body{margin:0;background:#f5f7fa;color:#101828;font-family:Inter,Arial,sans-serif}
.w{max-width:1180px;margin:auto;padding:34px 22px 64px}
.top{display:flex;justify-content:space-between;gap:24px;align-items:flex-start}
.k{font-size:12px;text-transform:uppercase;letter-spacing:.12em;color:#667085;font-weight:700}
h1{font-size:42px;line-height:1.08;margin:8px 0 10px}
h2{font-size:20px;margin:0 0 8px}
h3{font-size:15px;margin:0}
.sub{color:#667085;line-height:1.6;max-width:820px}
.card{background:#fff;border:1px solid #e4e7ec;border-radius:16px;padding:18px}
.sample{position:relative}
.sample summary{list-style:none;background:#fff;border:1px solid #d0d5dd;border-radius:10px;padding:11px 14px;font-size:13px;font-weight:700;cursor:pointer;white-space:nowrap}
.sample summary::-webkit-details-marker{display:none}
.sample-menu{position:absolute;right:0;top:48px;background:#fff;border:1px solid #e4e7ec;border-radius:12px;box-shadow:0 12px 32px rgba(16,24,40,.12);padding:8px;width:205px;z-index:10}
.sample-menu a{display:block;text-decoration:none;color:#101828;padding:10px;border-radius:8px;font-size:13px}
.sample-menu a:hover{background:#f2f4f7}
.setup{margin-top:28px}
.upload-card{padding:26px}
.upload-row{display:flex;gap:12px;flex-wrap:wrap;align-items:center;margin-top:18px}
.file-wrap{flex:1;min-width:280px;border:1px dashed #98a2b3;background:#fafafa;border-radius:12px;padding:10px}
input[type=file]{width:100%;font-size:13px}
button{background:#101828;color:white;border:0;padding:12px 17px;border-radius:10px;font-weight:700;cursor:pointer;font-size:14px}
button.secondary{background:#fff;color:#344054;border:1px solid #d0d5dd}
.error{color:#b42318;font-weight:700;margin-top:12px;font-size:13px}
.hidden{display:none!important}
.loading{margin:52px auto;max-width:560px;text-align:center}
.spinner{width:34px;height:34px;border:3px solid #e4e7ec;border-top-color:#101828;border-radius:50%;margin:0 auto 16px;animation:spin .8s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}
.result-head{display:flex;justify-content:space-between;gap:20px;align-items:center;margin:30px 0 16px}
.meta{color:#667085;font-size:13px}
.grid{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin:16px 0}
.metric{font-size:28px;font-weight:800;margin-top:6px}
.label{font-size:11px;color:#667085;text-transform:uppercase;letter-spacing:.05em}
.method{display:grid;grid-template-columns:1.25fr 1fr 1fr;gap:12px;margin:16px 0}
.method p,.guide p{font-size:13px;color:#667085;line-height:1.55;margin:7px 0 0}
.section{margin-top:16px}
.table-wrap{overflow:auto}
table{width:100%;border-collapse:collapse;font-size:13px}
th,td{padding:12px 10px;border-bottom:1px solid #eaecf0;text-align:left;vertical-align:top}
th{color:#667085;font-size:10px;text-transform:uppercase;letter-spacing:.05em}
.badge{display:inline-block;padding:4px 8px;border-radius:999px;font-size:11px;font-weight:700}
.High{background:#fee4e2;color:#b42318}.Medium{background:#fef0c7;color:#b54708}.Low{background:#dcfae6;color:#067647}
.signal-pill{display:inline-block;background:#f2f4f7;color:#344054;border-radius:999px;padding:4px 7px;font-size:11px;margin:0 4px 4px 0}
details.reason{min-width:280px}
details.reason summary{cursor:pointer;font-weight:700;color:#344054}
.reason-box{margin-top:10px;border-left:2px solid #d0d5dd;padding-left:12px}
.reason-summary{font-size:12px;line-height:1.55;color:#475467;margin-bottom:10px}
.signal{padding:10px 0;border-top:1px solid #f2f4f7}
.signal:first-of-type{border-top:0}
.signal-top{display:flex;justify-content:space-between;gap:12px}
.signal-name{font-size:12px;font-weight:800}
.signal-effect{font-size:11px;font-weight:700}
.raise{color:#b42318}.reduce{color:#067647}
.signal-meta{font-size:11px;color:#667085;margin:4px 0}
.signal-why{font-size:12px;color:#344054;line-height:1.45}
.signal-caution{font-size:11px;color:#98a2b3;line-height:1.45;margin-top:5px}
.bar-track{height:8px;background:#f2f4f7;border-radius:99px;overflow:hidden;min-width:180px}
.bar{height:100%;background:#667085;border-radius:99px}
.note{font-size:12px;color:#667085;line-height:1.55;margin-top:18px;padding:14px 16px;background:#f9fafb;border-radius:12px}
@media(max-width:950px){.grid{grid-template-columns:1fr 1fr}.method{grid-template-columns:1fr}.top,.result-head{align-items:flex-start}.top{flex-direction:column}.sample{align-self:flex-end;margin-top:-56px}}
@media(max-width:620px){h1{font-size:34px}.grid{grid-template-columns:1fr}.upload-row{align-items:stretch}.file-wrap{min-width:0}.sample{margin-top:0;align-self:flex-start}}
</style>
</head>
<body>
<div class="w">

<div id="preAnalysis">
  <div class="top">
    <div>
      <div class="k">Technyder / Workforce Analytics Demo</div>
      <h1>Employee Attrition Analyzer</h1>
      <div class="sub">
        Upload employee snapshots to measure historical churn and review current retention pressure.
        The analysis explains not only <strong>which signals moved</strong>, but also
        <strong>why the model raised them</strong> and how much each signal changed the score.
      </div>
    </div>

    <details class="sample" id="sampleMenu">
      <summary>Sample data ▾</summary>
      <div class="sample-menu">
        <a href="/api/employee-attrition/template.xlsx">Download Excel sample</a>
        <a href="/api/employee-attrition/template.csv">Download CSV sample</a>
      </div>
    </details>
  </div>

  <div class="setup">
    <div class="card upload-card">
      <div class="k">Start analysis</div>
      <h2 style="margin-top:7px">Upload CSV or Excel</h2>
      <div class="sub" style="font-size:13px">
        Use one row per employee snapshot. If <code>left_company</code> exists, the dashboard also
        calculates observed historical churn. Current active employees receive a retention-risk score.
      </div>
      <div class="upload-row">
        <div class="file-wrap"><input id="file" type="file" accept=".csv,.xlsx"></div>
        <button onclick="runAnalysis()">Run analysis</button>
      </div>
      <div id="error" class="error"></div>
    </div>
  </div>
</div>

<div id="loading" class="loading hidden">
  <div class="spinner"></div>
  <h2>Running workforce analysis</h2>
  <div class="sub" style="margin:auto">
    Validating employee rows, calculating attrition scores, comparing departments,
    and generating signal-level explanations.
  </div>
</div>

<div id="result" class="hidden">
  <div class="result-head">
    <div>
      <div class="k">Analysis complete</div>
      <h2 id="resultTitle" style="font-size:27px;margin-top:5px">Workforce retention analysis</h2>
      <div class="meta" id="fileMeta"></div>
    </div>
    <button class="secondary" onclick="resetAnalysis()">Analyze another file</button>
  </div>

  <div class="grid">
    <div class="card"><div class="label">Employees</div><div class="metric" id="employees">—</div></div>
    <div class="card"><div class="label">Observed churn</div><div class="metric" id="observed">—</div></div>
    <div class="card"><div class="label">Predicted attrition</div><div class="metric" id="predicted">—</div></div>
    <div class="card"><div class="label">High-risk active</div><div class="metric" id="highrisk">—</div></div>
    <div class="card"><div class="label">Avg engagement</div><div class="metric" id="engagement">—</div></div>
  </div>

  <div class="method">
    <div class="card guide">
      <div class="label">How to read predicted attrition</div>
      <h3 style="margin-top:7px">Portfolio pressure, not expected resignations</h3>
      <p>
        Predicted attrition is the average modeled risk across active employees.
        A 20% value means the current signal mix resembles a higher-pressure workforce state;
        it does not mean exactly 20% of employees will resign.
      </p>
    </div>
    <div class="card guide">
      <div class="label">Observed churn</div>
      <h3 style="margin-top:7px">Historical outcome</h3>
      <p>
        This KPI comes only from <code>left_company</code>. It describes the uploaded history and is
        intentionally kept separate from the current-risk estimate.
      </p>
    </div>
    <div class="card guide">
      <div class="label">Why a signal is raised</div>
      <h3 style="margin-top:7px">Value → benchmark → model effect</h3>
      <p>
        Each explanation shows the employee value, the reference point it crossed,
        and the approximate percentage-point effect on that employee's modeled risk.
      </p>
    </div>
  </div>

  <div class="card section">
    <div class="k">Priority review</div>
    <h2 style="margin-top:6px">Retention review queue</h2>
    <div class="sub" style="font-size:13px;margin-bottom:10px">
      Sorted by modeled attrition probability. Expand <strong>Why this risk?</strong> to see the conceptual reasoning.
    </div>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>Employee</th><th>Department</th><th>Role</th><th>Risk</th>
            <th>Probability</th><th>Primary signals</th><th>Model reasoning</th>
          </tr>
        </thead>
        <tbody id="rows"></tbody>
      </table>
    </div>
  </div>

  <div class="card section">
    <div class="k">Cohort view</div>
    <h2 style="margin-top:6px">Department attrition pressure</h2>
    <div class="sub" style="font-size:13px;margin-bottom:10px">
      Compare average modeled pressure across active employees. Use this view for workforce planning before drilling into individuals.
    </div>
    <div class="table-wrap">
      <table>
        <thead><tr><th>Department</th><th>Active employees</th><th>Predicted attrition</th><th>Relative pressure</th></tr></thead>
        <tbody id="departments"></tbody>
      </table>
    </div>
  </div>

  <div class="note" id="note"></div>
</div>
</div>

<script>
const pct=v=>v===null||v===undefined?'N/A':(v*100).toFixed(1)+'%';
const esc=value=>String(value??'').replace(/[&<>"']/g,char=>({
  '&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#039;'
}[char]));

function signalHtml(signal){
  const cls=signal.direction==='raises risk'?'raise':'reduce';
  const effect=(signal.model_effect_pp>=0?'+':'')+signal.model_effect_pp.toFixed(1)+' pp';
  return '<div class="signal">'+
    '<div class="signal-top"><div class="signal-name">'+esc(signal.signal)+'</div>'+
    '<div class="signal-effect '+cls+'">'+effect+'</div></div>'+
    '<div class="signal-meta"><strong>Observed:</strong> '+esc(signal.display_value)+
    ' · <strong>Reference:</strong> '+esc(signal.benchmark)+'</div>'+
    '<div class="signal-why">'+esc(signal.why)+'</div>'+
    '<div class="signal-caution">'+esc(signal.caution)+'</div>'+
  '</div>';
}

function employeeReasoning(employee){
  const signals=employee.top_signals.map(signalHtml).join('');
  return '<details class="reason"><summary>Why this risk?</summary>'+
    '<div class="reason-box"><div class="reason-summary">'+
      esc(employee.model_context.summary)+
    '</div>'+signals+'</div></details>';
}

async function runAnalysis(){
  const input=document.getElementById('file');
  const file=input.files[0];
  const error=document.getElementById('error');
  error.textContent='';

  if(!file){
    error.textContent='Choose a CSV or Excel file first.';
    return;
  }

  document.getElementById('preAnalysis').classList.add('hidden');
  document.getElementById('loading').classList.remove('hidden');

  const form=new FormData();
  form.append('file',file);

  try{
    const response=await fetch('/api/employee-attrition/upload',{
      method:'POST',
      body:form
    });
    const data=await response.json();

    if(!response.ok){
      throw new Error(
        typeof data.detail==='string'
          ? data.detail
          : JSON.stringify(data.detail)
      );
    }

    renderResults(data,file.name);
  }catch(err){
    document.getElementById('loading').classList.add('hidden');
    document.getElementById('preAnalysis').classList.remove('hidden');
    error.textContent=err.message;
  }
}

function renderResults(data,fileName){
  const k=data.kpis;

  document.getElementById('employees').textContent=k.unique_employees.toLocaleString();
  document.getElementById('observed').textContent=pct(k.observed_churn_rate);
  document.getElementById('predicted').textContent=pct(k.predicted_attrition_rate);
  document.getElementById('highrisk').textContent=k.high_risk_active_employees.toLocaleString();
  document.getElementById('engagement').textContent=k.average_engagement_score ?? 'N/A';

  document.getElementById('fileMeta').textContent=
    fileName+' · '+data.methodology.model_type+
    ' · High risk ≥ '+pct(data.methodology.high_risk_threshold);

  document.getElementById('rows').innerHTML=data.employees.slice(0,75).map(employee=>{
    const pills=employee.top_signals.slice(0,3).map(signal=>
      '<span class="signal-pill">'+esc(signal.signal)+'</span>'
    ).join('');

    return '<tr>'+
      '<td><strong>'+esc(employee.employee_id)+'</strong></td>'+
      '<td>'+esc(employee.department)+'</td>'+
      '<td>'+esc(employee.role_level)+'</td>'+
      '<td><span class="badge '+esc(employee.risk_band)+'">'+esc(employee.risk_band)+'</span></td>'+
      '<td><strong>'+pct(employee.attrition_probability)+'</strong><div class="meta">'+
        (employee.model_context.risk_above_baseline_pp>=0?'+':'')+
        employee.model_context.risk_above_baseline_pp.toFixed(1)+' pp vs baseline</div></td>'+
      '<td>'+pills+'</td>'+
      '<td>'+employeeReasoning(employee)+'</td>'+
    '</tr>';
  }).join('');

  const maxRate=Math.max(
    0.01,
    ...data.department_summary.map(item=>item.predicted_attrition_rate)
  );

  document.getElementById('departments').innerHTML=data.department_summary.map(department=>{
    const width=Math.max(3,(department.predicted_attrition_rate/maxRate)*100);
    return '<tr>'+
      '<td><strong>'+esc(department.department)+'</strong></td>'+
      '<td>'+department.active_employees+'</td>'+
      '<td>'+pct(department.predicted_attrition_rate)+'</td>'+
      '<td><div class="bar-track"><div class="bar" style="width:'+width.toFixed(1)+'%"></div></div></td>'+
    '</tr>';
  }).join('');

  document.getElementById('note').innerHTML=
    '<strong>Interpretation guardrail:</strong> '+esc(data.model_note)+
    '<br><br><strong>Model concept:</strong> '+esc(data.methodology.interpretation);

  document.getElementById('loading').classList.add('hidden');
  document.getElementById('result').classList.remove('hidden');
  window.scrollTo({top:0,behavior:'smooth'});
}

function resetAnalysis(){
  document.getElementById('result').classList.add('hidden');
  document.getElementById('file').value='';
  document.getElementById('error').textContent='';
  document.getElementById('preAnalysis').classList.remove('hidden');
  window.scrollTo({top:0,behavior:'smooth'});
}
</script>
</body>
</html>"""
